import json
import sys
import io
from pathlib import Path
from datetime import datetime

import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from core.scanner import run_scan

st.set_page_config(page_title="ApexRecon", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: 'Segoe UI', system-ui, sans-serif !important;
    background-color: #0d1117 !important;
    color: #e6edf3 !important;
    font-size: 15px !important;
}
.stApp { background-color: #0d1117 !important; }

section[data-testid="stSidebar"] {
    background-color: #0a0e14 !important;
    border-right: 1px solid #1e2432 !important;
    min-width: 240px !important;
    max-width: 240px !important;
}
section[data-testid="stSidebar"] * { color: #e6edf3 !important; font-size: 14px !important; }

.stTabs [data-baseweb="tab-list"] {
    background-color: #161b22 !important;
    border-radius: 12px !important;
    padding: 5px !important;
    gap: 4px !important;
    border-bottom: none !important;
}
.stTabs [data-baseweb="tab"] {
    background-color: transparent !important;
    border-radius: 8px !important;
    color: #6e7681 !important;
    font-size: 14px !important;
    font-weight: 500 !important;
    padding: 10px 22px !important;
    border: none !important;
}
.stTabs [aria-selected="true"] {
    background: #2f81f7 !important;
    color: #fff !important;
    font-weight: 700 !important;
    border-radius: 30px !important;
}
.stTabs [data-baseweb="tab-highlight"] { display: none !important; }
.stTabs [data-baseweb="tab-border"]    { display: none !important; }

div[data-testid="metric-container"] {
    background-color: #161b22 !important;
    border-radius: 14px !important;
    padding: 22px 24px !important;
    border: 1px solid #1e2432 !important;
}
div[data-testid="metric-container"] label {
    font-size: 13px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.6px !important;
    color: #8b949e !important;
    font-weight: 600 !important;
}
div[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-size: 34px !important;
    font-weight: 800 !important;
    color: #ffffff !important;
}

.stTextInput > div > div > input {
    background: #161b22 !important;
    border: 1px solid #1e2432 !important;
    border-radius: 8px !important;
    color: #e6edf3 !important;
    font-size: 14px !important;
    padding: 10px 14px !important;
}
.stTextInput > div > div > input:focus { border-color: #2f81f7 !important; }

.stButton > button {
    background: #2f81f7 !important;
    color: #fff !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    font-size: 14px !important;
    padding: 12px 24px !important;
    width: 100% !important;
}
.stButton > button:hover { background: #1a6cd4 !important; }

.stDownloadButton > button {
    background: #161b22 !important;
    color: #e6edf3 !important;
    border: 1px solid #1e2432 !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 14px !important;
    width: 100% !important;
    padding: 12px 0 !important;
}

.stDataFrame { background: #161b22 !important; border-radius: 10px !important; }
.stDataFrame * { color: #e6edf3 !important; font-size: 14px !important; }

div[data-testid="stExpander"] {
    background: #161b22 !important;
    border-radius: 10px !important;
    border: 1px solid #1e2432 !important;
}
div[data-testid="stExpander"] summary { font-size: 14px !important; color: #e6edf3 !important; }

h1, h2, h3, h4 { color: #ffffff !important; font-weight: 700 !important; }
p, li { color: #e6edf3 !important; font-size: 15px !important; }

.card {
    background: #161b22;
    border-radius: 14px;
    padding: 22px 24px;
    margin-bottom: 12px;
    border: 1px solid #1e2432;
}
.card-accent {
    background: linear-gradient(135deg, #2f81f7, #1a5fb4);
    border-radius: 14px;
    padding: 22px 24px;
}
.card-accent .lbl { font-size: 13px; color: rgba(255,255,255,0.7); text-transform: uppercase; letter-spacing: 0.6px; margin-bottom: 8px; font-weight: 600; }
.card-accent .val { font-size: 36px; font-weight: 800; color: #fff; line-height: 1; }
.card-accent .sub { font-size: 13px; color: rgba(255,255,255,0.6); margin-top: 6px; }

.flag-item {
    background: #0d1a2e;
    border-left: 4px solid #2f81f7;
    border-radius: 8px;
    padding: 14px 18px;
    margin-bottom: 8px;
    font-size: 14px;
    color: #a8d4ff;
    font-weight: 600;
}
.row-item {
    background: #161b22;
    border-radius: 10px;
    padding: 14px 18px;
    margin-bottom: 8px;
    font-size: 14px;
    color: #e6edf3;
    display: flex;
    justify-content: space-between;
    align-items: center;
    border: 1px solid #1e2432;
}
.tag {
    background: #1e2432;
    color: #8b949e;
    font-size: 13px;
    font-weight: 600;
    padding: 4px 12px;
    border-radius: 20px;
}
.tag-orange {
    background: #0d1f3c;
    color: #2f81f7;
    font-size: 13px;
    font-weight: 700;
    padding: 4px 12px;
    border-radius: 20px;
}
.section-lbl {
    font-size: 13px;
    font-weight: 700;
    color: #2f81f7;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin-bottom: 14px;
    margin-top: 4px;
}
.main .block-container {
    max-width: 100% !important;
    padding-left: 2.5rem !important;
    padding-right: 2.5rem !important;
    padding-top: 2rem !important;
}
hr { border-color: #1e2432 !important; }
</style>
""", unsafe_allow_html=True)


def dark_chart(fig):
    fig.update_layout(
        paper_bgcolor="#13131f", plot_bgcolor="#13131f",
        font_color="#c8c9d8", font_family="Segoe UI",
        margin=dict(l=10, r=10, t=36, b=10),
        legend=dict(bgcolor="#13131f"),
    )
    fig.update_xaxes(gridcolor="#1e1e2e", zerolinecolor="#1e1e2e")
    fig.update_yaxes(gridcolor="#1e1e2e", zerolinecolor="#1e1e2e")
    return fig


if "scan_results" not in st.session_state:
    st.session_state.scan_results = None


with st.sidebar:
    st.markdown("""
    <div style="padding-bottom:16px;border-bottom:1px solid #1a1a2a;margin-bottom:14px;">
      <div style="font-size:20px;font-weight:800;color:#fff;">Apex<span style="color:#2f81f7">Recon</span></div>
      <div style="font-size:12px;color:#6e7681;text-transform:uppercase;letter-spacing:1px;margin-top:3px;">Attack Surface Reconnaissance</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div style="font-size:13px;color:#8b949e;text-transform:uppercase;letter-spacing:0.8px;margin-bottom:8px;">Target Domain</div>', unsafe_allow_html=True)
    domain_input = st.text_input("", placeholder="example.com", label_visibility="collapsed")

    st.markdown('<div style="font-size:13px;color:#8b949e;text-transform:uppercase;letter-spacing:0.8px;margin-top:16px;margin-bottom:8px;">API Keys (optional)</div>', unsafe_allow_html=True)
    shodan_key = st.text_input("Shodan",     type="password", placeholder="Shodan API key",     label_visibility="collapsed")
    vt_key     = st.text_input("VirusTotal", type="password", placeholder="VirusTotal API key", label_visibility="collapsed")
    hibp_key   = st.text_input("HIBP",       type="password", placeholder="HIBP API key",       label_visibility="collapsed")

    st.markdown("<br>", unsafe_allow_html=True)
    run_btn = st.button("Run Scan")

    st.markdown('<div style="margin-top:32px;font-size:12px;color:#3a4050;text-transform:uppercase;letter-spacing:0.8px;">ApexRecon v1.0</div>', unsafe_allow_html=True)


if run_btn and domain_input:
    with st.spinner(f"Scanning {domain_input}..."):
        st.session_state.scan_results = run_scan(domain_input.strip(), {
            "shodan_key": shodan_key or None,
            "vt_key":     vt_key     or None,
            "hibp_key":   hibp_key   or None,
        })

r = st.session_state.scan_results

st.markdown(f"""
<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:20px;">
  <div>
    <h2 style="margin:0;font-size:22px;font-weight:700;color:#fff;">{"Scan Results - " + r["target"] if r else "ApexRecon"}</h2>
    <p style="color:#3a3b55;font-size:12px;margin-top:4px;">
      {"Scan completed " + r.get("finished_at","")[:19].replace("T"," ") + " in " + str(r.get("duration_s","?")) + "s" if r else "Enter a domain in the sidebar to start a scan."}
    </p>
  </div>
  <div style="background:#13131f;border-radius:20px;padding:7px 14px;font-size:11px;color:#2f81f7;">
    {"HIGH RISK" if r and r.get("summary",{}).get("risk_level")=="high" else "MEDIUM RISK" if r and r.get("summary",{}).get("risk_level")=="medium" else "LOW RISK" if r else "IDLE"}
  </div>
</div>
""", unsafe_allow_html=True)

if not r:
    st.info("No scan results yet. Enter a domain and click Run Scan.")
    st.stop()

mods    = r.get("modules", {})
summary = r.get("summary", {})
stats   = summary.get("stats", {})
flags   = summary.get("flags", [])

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "Overview", "DNS & Subdomains", "Infrastructure", "Security Audit", "Threat Intel", "Report"
])


with tab1:
    c0, c1, c2, c3, c4 = st.columns(5)
    with c0:
        st.markdown(f'<div class="card-accent"><div class="lbl">Risk Flags</div><div class="val">{summary.get("flag_count",0)}</div><div class="sub">{summary.get("risk_level","?").upper()} RISK</div></div>', unsafe_allow_html=True)
    c1.metric("Subdomains (DNS)",  stats.get("subdomains_dns", 0))
    c2.metric("Subdomains (Cert)", stats.get("subdomains_cert", 0))
    c3.metric("Open Ports",        stats.get("open_ports", 0))
    c4.metric("Headers Grade",     stats.get("headers_grade", "N/A"))

    st.markdown("<br>", unsafe_allow_html=True)
    col_flags, col_whois = st.columns([1.4, 1])

    with col_flags:
        st.markdown('<div class="section-lbl">Risk Flags</div>', unsafe_allow_html=True)
        if flags:
            for f in flags:
                st.markdown(f'<div class="flag-item">{f}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="card" style="color:#4a4b6a;font-size:13px;">No flags raised.</div>', unsafe_allow_html=True)

    with col_whois:
        st.markdown('<div class="section-lbl">WHOIS Summary</div>', unsafe_allow_html=True)
        wd = mods.get("whois", {}).get("data", {})
        for label, key in [("Registrar","registrar"),("Org","registrant_org"),("Country","registrant_country"),("Created","creation_date"),("Expires","expiry_date"),("DNSSEC","dnssec")]:
            val = wd.get(key)
            if val:
                v = val if not isinstance(val, list) else val[0]
                st.markdown(f'<div class="row-item"><span style="color:#8b949e;font-size:14px;font-weight:600;">{label}</span><span style="font-size:12px;">{v}</span></div>', unsafe_allow_html=True)
        if wd.get("domain_age_days"):
            st.markdown(f'<div class="row-item"><span style="color:#8b949e;font-size:14px;font-weight:600;">Domain Age</span><span style="font-size:12px;">{wd["domain_age_days"]} days</span></div>', unsafe_allow_html=True)

    geo = mods.get("geoip", {}).get("data", {})
    if geo.get("latitude") and geo.get("longitude"):
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="section-lbl">Server Location</div>', unsafe_allow_html=True)

        map_df = pd.DataFrame([{"lat": geo["latitude"], "lon": geo["longitude"]}])
        st.map(map_df, zoom=4)
        st.markdown(f'<div class="row-item" style="margin-top:8px;"><span style="color:#8b949e;font-size:14px;font-weight:600;">Location</span><span style="font-size:12px;">{geo.get("city","?")}, {geo.get("country","?")}</span></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="row-item"><span style="color:#8b949e;font-size:14px;font-weight:600;">ASN / Org</span><span style="font-size:12px;">{geo.get("org","?")}</span></div>', unsafe_allow_html=True)


with tab2:
    st.markdown('<div class="section-lbl">DNS Records</div>', unsafe_allow_html=True)
    dns_data = mods.get("dns", {}).get("data", {})
    records  = dns_data.get("records", {})
    if records:
        for rtype, vals in records.items():
            if vals:
                with st.expander(f"{rtype} records ({len(vals)})"):
                    for v in vals:
                        st.markdown(f'<div class="row-item">{v}</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col_dns, col_cert = st.columns(2)

    with col_dns:
        st.markdown('<div class="section-lbl">Subdomains via DNS Brute-force</div>', unsafe_allow_html=True)
        dns_subs = dns_data.get("subdomains", [])
        if dns_subs:
            for sub in dns_subs:
                st.markdown(f'<div class="row-item"><span>{sub["subdomain"]}</span><span class="tag">{sub["ip"]}</span></div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="card" style="color:#4a4b6a;font-size:13px;">No subdomains resolved.</div>', unsafe_allow_html=True)

    with col_cert:
        st.markdown('<div class="section-lbl">Subdomains via Certificate Transparency</div>', unsafe_allow_html=True)
        cert_data = mods.get("certlog", {}).get("data", {})
        cert_subs = cert_data.get("subdomains", [])
        if cert_subs:
            for sub in cert_subs[:40]:
                st.markdown(f'<div class="row-item">{sub}</div>', unsafe_allow_html=True)
            if len(cert_subs) > 40:
                st.markdown(f'<div style="font-size:11px;color:#4a4b6a;margin-top:6px;">+ {len(cert_subs)-40} more</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="card" style="color:#4a4b6a;font-size:13px;">No results from crt.sh.</div>', unsafe_allow_html=True)

    all_subs = list(set([s["subdomain"] for s in dns_subs] + cert_subs))
    if all_subs:
        st.markdown(f'<div class="section-lbl" style="margin-top:16px;">Total unique subdomains: {len(all_subs)}</div>', unsafe_allow_html=True)
        certs = cert_data.get("certificates", [])
        if certs:
            st.markdown('<div class="section-lbl" style="margin-top:12px;">TLS Certificates</div>', unsafe_allow_html=True)
            cert_df = pd.DataFrame(certs)[["common_name","issuer","not_before","not_after"]].head(20)
            st.dataframe(cert_df, use_container_width=True)


with tab3:
    shodan_data = mods.get("shodan", {}).get("data", {})
    geo_data    = mods.get("geoip",  {}).get("data", {})

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("IP Address", shodan_data.get("ip") or geo_data.get("ip", "N/A"))
    col2.metric("Open Ports", len(shodan_data.get("open_ports", [])))
    col3.metric("CVEs Found", len(shodan_data.get("vulns", [])))
    col4.metric("OS",         shodan_data.get("os") or "Unknown")

    st.markdown("<br>", unsafe_allow_html=True)
    col_ports, col_info = st.columns([1.5, 1])

    with col_ports:
        st.markdown('<div class="section-lbl">Open Ports and Services</div>', unsafe_allow_html=True)
        services = shodan_data.get("services", [])
        if services:
            for svc in services:
                product = svc.get("product") or "unknown"
                version = svc.get("version") or ""
                st.markdown(f'<div class="row-item"><span><b style="color:#fff;">{svc["port"]}/{svc["transport"]}</b> <span style="color:#4a4b6a;">{product} {version}</span></span><span class="tag-orange">open</span></div>', unsafe_allow_html=True)
            ports  = [str(s["port"]) for s in services]
            fig = px.bar(x=ports, y=[1]*len(ports), title="Open Ports", color_discrete_sequence=["#2f81f7"])
            fig.update_layout(xaxis_title="Port", yaxis_title="", showlegend=False)
            st.plotly_chart(dark_chart(fig), use_container_width=True)
        elif mods.get("shodan", {}).get("status") == "skipped":
            st.markdown('<div class="card" style="color:#4a4b6a;font-size:13px;">Add a Shodan API key to see open ports and services.</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="card" style="color:#4a4b6a;font-size:13px;">Host not indexed in Shodan.</div>', unsafe_allow_html=True)

    with col_info:
        st.markdown('<div class="section-lbl">Host Info</div>', unsafe_allow_html=True)
        for label, val in [("Org", shodan_data.get("org") or geo_data.get("org")), ("ISP", shodan_data.get("isp")), ("Country", shodan_data.get("country") or geo_data.get("country")), ("City", shodan_data.get("city") or geo_data.get("city")), ("ASN", geo_data.get("asn")), ("Timezone", geo_data.get("timezone"))]:
            if val:
                st.markdown(f'<div class="row-item"><span style="color:#8b949e;font-size:14px;font-weight:600;">{label}</span><span style="font-size:12px;">{val}</span></div>', unsafe_allow_html=True)
        vulns = shodan_data.get("vulns", [])
        if vulns:
            st.markdown('<div class="section-lbl" style="margin-top:16px;">CVEs</div>', unsafe_allow_html=True)
            for v in vulns:
                st.markdown(f'<div class="flag-item">{v}</div>', unsafe_allow_html=True)


with tab4:
    headers_data = mods.get("headers", {}).get("data", {})
    grade   = headers_data.get("grade", "N/A")
    present = headers_data.get("present_headers", {})
    missing = headers_data.get("missing_headers", [])
    hflags  = headers_data.get("flags", [])

    col1, col2, col3 = st.columns(3)
    col1.metric("Security Grade",  grade)
    col2.metric("Headers Present", len(present))
    col3.metric("Headers Missing", len(missing))

    st.markdown("<br>", unsafe_allow_html=True)
    col_pass, col_fail = st.columns(2)

    with col_pass:
        st.markdown('<div class="section-lbl">Present</div>', unsafe_allow_html=True)
        for h, v in present.items():
            val = v[:60] + "..." if isinstance(v, str) and len(v) > 60 else v
            st.markdown(f'<div class="row-item"><span style="color:#1D9E75;font-size:12px;">check {h}</span><span class="tag" style="max-width:160px;overflow:hidden;text-overflow:ellipsis;">{val}</span></div>', unsafe_allow_html=True)

    with col_fail:
        st.markdown('<div class="section-lbl">Missing</div>', unsafe_allow_html=True)
        for h in missing:
            st.markdown(f'<div class="row-item"><span style="color:#2f81f7;font-size:12px;">x {h}</span></div>', unsafe_allow_html=True)

    if hflags:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="section-lbl">Exposure Flags</div>', unsafe_allow_html=True)
        for f in hflags:
            st.markdown(f'<div class="flag-item">{f}</div>', unsafe_allow_html=True)

    names  = list(present.keys()) + missing
    labels = ["Present"] * len(present) + ["Missing"] * len(missing)
    if names:
        fig = px.pie(names=names, values=[1]*len(names), color=labels,
            color_discrete_map={"Present": "#1D9E75", "Missing": "#2f81f7"},
            title="Security Headers Coverage")
        st.plotly_chart(dark_chart(fig), use_container_width=True)


with tab5:
    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<div class="section-lbl">VirusTotal</div>', unsafe_allow_html=True)
        vt = mods.get("virustotal", {})
        if vt.get("status") == "skipped":
            st.markdown('<div class="card" style="color:#4a4b6a;font-size:13px;">Add a VirusTotal API key to see threat intelligence.</div>', unsafe_allow_html=True)
        else:
            vd = vt.get("data", {})
            for label, key in [("Reputation","reputation"),("Malicious","malicious"),("Suspicious","suspicious"),("Harmless","harmless"),("Registrar","registrar")]:
                val = vd.get(key)
                if val is not None:
                    st.markdown(f'<div class="row-item"><span style="color:#8b949e;font-size:14px;font-weight:600;">{label}</span><span style="font-size:12px;">{val}</span></div>', unsafe_allow_html=True)
            if vd.get("flag"):
                st.markdown(f'<div class="flag-item" style="margin-top:10px;">{vd["flag"]}</div>', unsafe_allow_html=True)
            stats_vt = {"Malicious": vd.get("malicious",0), "Suspicious": vd.get("suspicious",0), "Harmless": vd.get("harmless",0), "Undetected": vd.get("undetected",0)}
            fig = px.bar(x=list(stats_vt.keys()), y=list(stats_vt.values()), title="Vendor Analysis Breakdown",
                color=list(stats_vt.keys()), color_discrete_sequence=["#c0392b","#2f81f7","#1D9E75","#4a4b6a"])
            fig.update_layout(showlegend=False, xaxis_title="", yaxis_title="Vendors")
            st.plotly_chart(dark_chart(fig), use_container_width=True)

    with col2:
        st.markdown('<div class="section-lbl">Have I Been Pwned</div>', unsafe_allow_html=True)
        hibp = mods.get("hibp", {})
        if hibp.get("status") == "skipped":
            st.markdown(f'<div class="card" style="color:#4a4b6a;font-size:13px;">{hibp.get("reason","no api key")}.</div>', unsafe_allow_html=True)
        else:
            hd = hibp.get("data", {})
            bc = hd.get("breached_count", 0)
            st.markdown(f'<div class="row-item"><span style="color:#8b949e;font-size:14px;font-weight:600;">Emails checked</span><span>{hd.get("total_checked",0)}</span></div>', unsafe_allow_html=True)
            st.markdown(f'<div class="row-item"><span style="color:#8b949e;font-size:14px;font-weight:600;">Breached accounts</span><span style="color:{"#c0392b" if bc > 0 else "#1D9E75"};">{bc}</span></div>', unsafe_allow_html=True)
            for check in hd.get("results", []):
                if check.get("breached"):
                    with st.expander(f"{check['email']} - {check['count']} breach(es)"):
                        for b in check["breaches"]:
                            st.markdown(f'<div class="row-item"><span><b style="color:#fff;">{b["name"]}</b> <span style="color:#4a4b6a;">({b["breach_date"]})</span></span><span class="tag">{b["domain"]}</span></div>', unsafe_allow_html=True)


with tab6:
    from fpdf import FPDF
    import openpyxl

    st.markdown('<div class="section-lbl">Scan Summary</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="row-item"><span style="color:#4a4b6a;">Target</span><span>{r["target"]}</span></div>
    <div class="row-item"><span style="color:#4a4b6a;">Scanned</span><span>{r.get("started_at","")[:19].replace("T"," ")}</span></div>
    <div class="row-item"><span style="color:#4a4b6a;">Duration</span><span>{r.get("duration_s","?")}s</span></div>
    <div class="row-item"><span style="color:#4a4b6a;">Risk Level</span><span style="color:#2f81f7;font-weight:600;">{summary.get("risk_level","?").upper()}</span></div>
    <div class="row-item"><span style="color:#4a4b6a;">Flags Raised</span><span>{summary.get("flag_count",0)}</span></div>
    """, unsafe_allow_html=True)

    if flags:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="section-lbl">Flags</div>', unsafe_allow_html=True)
        for f in flags:
            st.markdown(f'<div class="flag-item">{f}</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-lbl">Export Report</div>', unsafe_allow_html=True)

    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_target = r["target"].replace(".", "_")
    dns_subs    = mods.get("dns",{}).get("data",{}).get("subdomains",[])
    cert_subs   = mods.get("certlog",{}).get("data",{}).get("subdomains",[])

    c1, c2, c3 = st.columns(3)

    with c1:
        st.download_button(
            "Download JSON",
            json.dumps(r, indent=2, default=str),
            file_name=f"apexrecon_{safe_target}_{ts}.json",
            mime="application/json",
            use_container_width=True,
        )

    with c2:
        wb         = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Field", "Value"])
        ws_summary.append(["Target",     r["target"]])
        ws_summary.append(["Scanned",    r.get("started_at","")[:19]])
        ws_summary.append(["Duration",   f"{r.get('duration_s','?')}s"])
        ws_summary.append(["Risk Level", summary.get("risk_level","?").upper()])
        ws_summary.append(["Flags",      summary.get("flag_count",0)])
        ws_summary.append([])
        ws_summary.append(["Flags Raised"])
        for f in flags:
            ws_summary.append(["", f])

        ws_dns = wb.create_sheet("DNS and Subdomains")
        ws_dns.append(["Subdomain", "IP", "Source"])
        for s in dns_subs:
            ws_dns.append([s.get("subdomain",""), s.get("ip",""), "DNS brute-force"])
        for s in cert_subs:
            ws_dns.append([s, "", "Certificate transparency"])

        ws_ports = wb.create_sheet("Open Ports")
        ws_ports.append(["Port", "Transport", "Product", "Version"])
        for svc in mods.get("shodan",{}).get("data",{}).get("services",[]):
            ws_ports.append([svc.get("port"), svc.get("transport"), svc.get("product"), svc.get("version")])

        ws_headers = wb.create_sheet("Headers Audit")
        ws_headers.append(["Header", "Status", "Value"])
        for h, v in mods.get("headers",{}).get("data",{}).get("present_headers",{}).items():
            ws_headers.append([h, "Present", str(v)[:100]])
        for h in mods.get("headers",{}).get("data",{}).get("missing_headers",[]):
            ws_headers.append([h, "Missing", ""])

        xl_buf = io.BytesIO()
        wb.save(xl_buf)
        xl_buf.seek(0)
        st.download_button(
            "Download Excel",
            xl_buf.getvalue(),
            file_name=f"apexrecon_{safe_target}_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with c3:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "ApexRecon - Scan Report", ln=True)
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 6, f"Target: {r['target']}", ln=True)
        pdf.cell(0, 6, f"Scanned: {r.get('started_at','')[:19]}", ln=True)
        pdf.cell(0, 6, f"Duration: {r.get('duration_s','?')}s", ln=True)
        pdf.cell(0, 6, f"Risk Level: {summary.get('risk_level','?').upper()}", ln=True)
        pdf.ln(4)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Flags Raised", ln=True)
        pdf.set_font("Helvetica", size=10)
        if flags:
            for f in flags:
                pdf.multi_cell(0, 6, f"- {f}")
        else:
            pdf.cell(0, 6, "No flags raised.", ln=True)
        pdf.ln(4)

        whois_d = mods.get("whois",{}).get("data",{})
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "WHOIS", ln=True)
        pdf.set_font("Helvetica", size=10)
        for k, v in whois_d.items():
            if v and k not in ("flag",):
                safe_val = str(v)[:80].encode("latin-1", errors="replace").decode("latin-1")
                pdf.cell(0, 6, f"{k}: {safe_val}", ln=True)
        pdf.ln(4)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Subdomains", ln=True)
        pdf.set_font("Helvetica", size=10)
        all_subs = [s["subdomain"] for s in dns_subs] + list(cert_subs)
        for s in sorted(set(all_subs))[:40]:
            safe_s = s.encode("latin-1", errors="replace").decode("latin-1")
            pdf.cell(0, 5, safe_s, ln=True)
        if len(all_subs) > 40:
            pdf.cell(0, 5, f"... and {len(all_subs)-40} more", ln=True)
        pdf.ln(4)

        headers_d = mods.get("headers",{}).get("data",{})
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, f"Security Headers - Grade {headers_d.get('grade','N/A')}", ln=True)
        pdf.set_font("Helvetica", size=10)
        for h in headers_d.get("missing_headers",[]):
            pdf.cell(0, 5, f"Missing: {h}", ln=True)

        pdf_bytes = pdf.output()
        st.download_button(
            "Download PDF",
            bytes(pdf_bytes),
            file_name=f"apexrecon_{safe_target}_{ts}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
